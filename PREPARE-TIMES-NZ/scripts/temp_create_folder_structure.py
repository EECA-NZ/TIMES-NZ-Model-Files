import sys
import os
from openpyxl import load_workbook
from pathlib import Path


current_dir = Path(__file__).resolve().parent
sys.path.append(os.path.join(current_dir, "..", "library"))
from config import DATA_INTERMEDIATE, TIMES_INPUT_FILES

source_folder = TIMES_INPUT_FILES
destination_folder = Path(f"{DATA_INTERMEDIATE}/automatic folder production")


def create_excel_folder_structure(source_folder, destination_folder):
    """
    Creates a folder structure based on Excel files and their sheets.
    
    Args:
        source_folder (str): Path to folder containing Excel files
        destination_folder (str): Path where new folder structure will be created
    """
    # Create the destination folder if it doesn't exist
    os.makedirs(destination_folder, exist_ok=True)
    
    # Get all Excel files in the source folder
    excel_files = [f for f in os.listdir(source_folder) 
                  if f.endswith(('.xlsx', '.xls', '.xlsm'))]
    
    # Process each Excel file
    for excel_file in excel_files:
        # Create folder for the Excel file
        excel_folder_name = os.path.splitext(excel_file)[0]  # Remove extension
        excel_folder_path = os.path.join(destination_folder, excel_folder_name)
        os.makedirs(excel_folder_path, exist_ok=True)
        
        # Load the workbook
        workbook_path = os.path.join(source_folder, excel_file)
        try:
            workbook = load_workbook(workbook_path, read_only=True)
            
            # Create folder for each sheet
            for sheet_name in workbook.sheetnames:
                # Clean sheet name to be valid folder name
                clean_sheet_name = "".join(c for c in sheet_name 
                                         if c not in r'\/:*?"<>|')
                sheet_folder_path = os.path.join(excel_folder_path, clean_sheet_name)
                os.makedirs(sheet_folder_path, exist_ok=True)
                
            workbook.close()
            
        except Exception as e:
            print(f"Error processing {excel_file}: {str(e)}")

# Example usage

create_excel_folder_structure(source_folder, destination_folder)