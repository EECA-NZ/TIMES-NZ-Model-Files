
import sys 
import os 
from openpyxl import Workbook, load_workbook
from ast import literal_eval
import pandas as pd 
import numpy as np
import string
import logging


# bit of a mess will need to come back to this one 

# contain all functions to read the metadata file in data_intermediate/config 
# and use this to: 

# 1: create all the workbooks and their relevant sheets 
# 2: Identify all the tags we're going to write and their metadata (book, sheet, tag, uc_sets, maybe more) 
# 3: get all the dataframes from the original toml files 

#                   Hey I just had a crzy though - do we ust make a bunch of csvs for the toml data? 
#                   honestly nah just leave them as dicts in the normalised files 

# 5

# get custom locations
logging.basicConfig(level=logging.INFO)

from filepaths import DATA_INTERMEDIATE, OUTPUT_LOCATION



def get_csv_data(file_location):           
    # must read as string in order to pull through full precision
    df = pd.read_csv(file_location, dtype = str)
    return df



def dict_to_dataframe(dict):
    # takes a single dictionary from our tomls and creates a dataframe
    # this is only used for the direct toml data
    df_parts = []
    for key, values in dict.items():
        # Create DataFrame with explicit index
        if isinstance(values, (str, int, float, bool)) or values is None:
            # For scalar values, we need to wrap in a list and provide an index
            df_part = pd.DataFrame({key: [values]})
        else:
            # For dictionary or other iterable values
            df_part = pd.DataFrame({key: values})
        df_parts.append(df_part)
    
    # Concatenate all DataFrames
    if df_parts:
        df = pd.concat(df_parts, axis=1)
        df = df.fillna("")
    else:
        # Handle empty dictionary case
        df = pd.DataFrame()
    
    return df


def strip_headers_from_tiny_df(df):
    """    
    We need special handling for some tables with one value
    XL2TIMES outputs these with a header called VALUE, but we need the value to be in the header column (with no data underneath)
    This function replaces the header with the value for specific tables 

    """
    df = df.T  # Transpose the df
    df.columns = [df.iloc[0][0]]  # Set the column name to the value
    df = df.iloc[0:0]  # Remove all rows 
    return df




def test_if_toml_location(string):
    return string.endswith('.toml')



def create_empty_workbook(book_name, sheets):
    # This function creates the workbook with empty sheets
    # Later, data is appended to these sheets by overlay.        
    book_location = f"{OUTPUT_LOCATION}/{book_name}.xlsx"
    book_directory = os.path.dirname(book_location)

    # create the folder if needed
    os.makedirs(book_directory, exist_ok=True)

    # create workbook:
    wb = Workbook()

    # remove the default sheet which gets activated
    wb.remove(wb.active)

    # add each sheet 
    for sheet in sheets: 
        wb.create_sheet(sheet)

    # save 
    wb.save(book_location)


def write_data(df, book_name, sheet_name, tag, uc_set, startrow=0):        
    

        
    # Load existing workbook
    book_location = f"{OUTPUT_LOCATION}/{book_name}.xlsx"
    book = load_workbook(book_location)
    sheet = book[sheet_name]

    


    # Get uc_set length and adjust startrow if needed


    uc_set_length = len(uc_set)
    if uc_set_length > 0:
        print("uc_sets detected")
        startrow += uc_set_length-1
    else: 
        print("no uc_sets detected ")
    
    # Write the header row
    for col_idx, column_name in enumerate(df.columns, 1):
        sheet.cell(row=startrow + 2, column=col_idx, value=column_name)
    
    # Write the data
    for row_idx, row in enumerate(df.values, startrow + 3):
        for col_idx, value in enumerate(row, 1):
            sheet.cell(row=row_idx, column=col_idx, value=value)
    
    # Write the tag
    tag_row = startrow + 1
    sheet.cell(row=tag_row, column=1, value=tag)
    
    # Add UC_Set tags if needed
    if uc_set_length > 0:
        for n in range(uc_set_length):
            uc_set_tag_row = startrow - n + 1
            key = list(uc_set.keys())[n]
            value = uc_set[key]
            sheet.cell(row=uc_set_tag_row, column=2, value=f"~UC_Sets: {key}: {value}")
    
    # Save the workbook
    book.save(book_location)


# probably won't use the below two functions but will keep these to farm for parts and delete when finished 
        
def write_all_tags_to_sheet(book_name, sheet_name):

    # The sheets with multiple tags need to be stacked up.
    # Each table we write will need to have its row number saved in here so other tables can move down. 
       
    tag_list = get_tags_for_sheet(book_name, sheet_name)
    sheet_folder = f"{INPUT_LOCATION}/{book_name}/{sheet_name}"

    # we start from the first row and will move down as needed
    startrow = 0

    for tag_name in tag_list:
        csv_files = return_csvs_in_folder(f"{sheet_folder}/{tag_name}")

        for csv_name in csv_files:
            # read the data 
            df = get_csv_data(book_name, sheet_name, tag_name, csv_name)
            # include the uc_set if needed (this comes up null otherwise)
            uc_set = get_uc_sets(book_name, sheet_name, tag_name, csv_name)

            # Patch for small files
            # TO-DO: automate handling of these rather than hardcoding which tables receive this treatment.
            # This currently covers all of TIMES-NZ but better to be flexible to future changes if needed.

            if (book_name == "SysSettings"
                    and sheet_name == "TimePeriods" 
                    and tag_name in ["StartYear", "ActivePDef"]):
                df = strip_headers_from_tiny_df(df)

            # create the tag (this also returns the colons where necessary)
            write_data(df, book_name, sheet_name, tag_name, uc_set, startrow = startrow)
            # measure the length (row count), adding extra space for additional uc_sets if needed, so the next table has space
            df_row_count= len(df) + len(uc_set)
            # add the dataframe rows to our start row index so we can keep going without overwriting
            # and additional rows for a healthy gap.            
            startrow += df_row_count + 3     
        
def write_workbook(book_name):
    logging.info(f"Creating {book_name}.xlsx:")
    sheets = get_sheets_for_book(book_name)
    # create structure, overwriting everything already there
    create_empty_workbook(book_name, sheets, suffix = "")

    for sheet in sheets: 
        # Verbose printing
        logging.info(f"     - Sheet: '{sheet}'")
        # the workbook exists now we write each tag set to each sheet 
        write_all_tags_to_sheet(book_name, sheet_name = sheet)


# test_data = toml_test["TimeSlices"]