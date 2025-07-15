"""
# bit of a mess will need to come back to this one

# contain all functions to read the metadata file in data_intermediate/config
# and use this to:

# 1: create all the workbooks and their relevant sheets
# 2: Identify all the tags we're going to write and their metadata
# (book, sheet, tag, uc_sets, maybe more)
# 3: get all the dataframes from the original toml files



"""

# Libraries ------------------------------------------------------------------
import os

import pandas as pd
from openpyxl import Workbook, load_workbook
from prepare_times_nz.filepaths import OUTPUT_LOCATION
from prepare_times_nz.logger_setup import logger


# Functions ------------------------------------------------------------------
def get_csv_data(file_location):
    """
    Returns the datafrom from a file_location,
    specifically we are ensuring that all data is read as strings
    This ensures full precision
    """
    df = pd.read_csv(file_location, dtype=str)
    return df


def dict_to_dataframe(data_dict):
    """
    takes a single dictionary from our tomls and creates a dataframe
    this is only used for the direct toml data
    (ie: dataframes entered into the toml files directly are converted here)
    """

    df_parts = []
    for key, values in data_dict.items():
        # Create DataFrame with explicit index
        if isinstance(values, (str, int, float, bool)) or values is None:
            # For scalar values, we need to wrap in a list and provide an index
            df_part = pd.DataFrame({key: [values]})
        else:
            # For dictionary or other iterable values
            df_part = pd.DataFrame({key: values})
        df_parts.append(df_part)

    # Concatenate all DataFrames
    if df_parts:
        df = pd.concat(df_parts, axis=1)
        df = df.fillna("")
    else:
        # Handle empty dictionary case
        df = pd.DataFrame()

    return df


def strip_headers_from_tiny_df(df):
    """
    We need special handling for some tables with one value
    XL2TIMES outputs these with a header called VALUE,
    but we need the value to be in the header column (with no data underneath)
    This function replaces the header with the value for specific tables

    """
    df = df.T  # Transpose the df
    df.columns = [df.iloc[0][0]]  # Set the column name to the value
    df = df.iloc[0:0]  # Remove all rows
    return df


def test_if_toml_location(string):
    """
    Return True if filename ends with .toml else False
    Mostly just wrapping the function in a more obvious name for its purpose

    """
    return string.endswith(".toml")


def create_empty_workbook(book_name, sheets):
    """
    Creates a workbook with empty sheets
    THe sheets will have data appended via excel overlay,
    but the sheets are required to exist first for that to work.

    Takes a book name and list of sheets, then saves the empty book
    in OUTPUT_LOCATION

    """
    # This function creates the workbook with empty sheets
    # Later, data is appended to these sheets by overlay.
    book_location = f"{OUTPUT_LOCATION}/{book_name}.xlsx"
    book_directory = os.path.dirname(book_location)

    # create the folder if needed
    os.makedirs(book_directory, exist_ok=True)

    # create workbook:
    wb = Workbook()

    # remove the default sheet which gets activated
    wb.remove(wb.active)

    # add each sheet
    for sheet in sheets:
        wb.create_sheet(sheet)

    # save
    wb.save(book_location)


def write_data(df, book_name, sheet_name, tag, uc_set, startrow=0):
    """
    Writes data to the existing book

    Based on the input dataframe, the excel workbook is updated based on:

      The book's name
      THe Sheet name,
      The required input tag,
      Any uc_sets if relevant
      THe start row to begin adding data to

    The startrow input is a required input because each new tag added to a sheet
    needs to increase the start row so our tables don't overwrite each other.

    THis allows us to print many tagged tables to a single worksheet

    """

    # Ideally, this function should be broken down into smaller pieces
    # to reduce the required arguments

    # pylint: disable=too-many-positional-arguments
    # pylint: disable=too-many-arguments
    # pylint: disable=too-many-locals

    # Load existing workbook
    book = load_workbook(f"{OUTPUT_LOCATION}/{book_name}.xlsx")
    sheet = book[sheet_name]

    # Get uc_set length and adjust startrow if needed
    uc_set_length = len(uc_set)
    if uc_set_length > 0:
        logger.debug("uc_sets detected")
        startrow += uc_set_length - 1
    else:
        logger.debug("no uc_sets detected ")

    # Write the header row
    for col_idx, column_name in enumerate(df.columns, 1):
        sheet.cell(row=startrow + 2, column=col_idx, value=column_name)

    # Write the data
    for row_idx, row in enumerate(df.values, startrow + 3):
        for col_idx, value in enumerate(row, 1):
            sheet.cell(row=row_idx, column=col_idx, value=value)

    # Write the tag
    tag_row = startrow + 1
    sheet.cell(row=tag_row, column=1, value=tag)

    # Add UC_Set tags if needed
    if uc_set_length > 0:
        for n in range(uc_set_length):
            uc_set_tag_row = startrow - n + 1
            key = list(uc_set.keys())[n]
            value = uc_set[key]
            sheet.cell(row=uc_set_tag_row, column=2, value=f"~UC_Sets: {key}: {value}")

    # Save the workbook
    book.save(f"{OUTPUT_LOCATION}/{book_name}.xlsx")
