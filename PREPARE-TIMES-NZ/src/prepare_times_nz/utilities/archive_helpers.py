"""
This module contains helper functions for working with Excel files and CSV data.
"""

import os
import string
from ast import literal_eval

import pandas as pd
from openpyxl import Workbook, load_workbook
from prepare_times_nz.utilities.filepaths import DATA_INTERMEDIATE, OUTPUT_LOCATION
from prepare_times_nz.utilities.logger_setup import logger

INPUT_LOCATION = DATA_INTERMEDIATE


def get_csv_data(book_name, sheet_name, tag_name, csv_name):
    """
    Reads a CSV file from a structured directory
         and returns its contents as a DataFrame.

    The file path is:
        {INPUT_LOCATION}/{book_name}/{sheet_name}/{tag_name}/{csv_name}.csv

    All columns are read as strings to preserve precision.

    Args:
        book_name (str): Book directory name.
        sheet_name (str): Sheet directory name.
        tag_name (str): Tag directory name.
        csv_name (str): CSV file name (without '.csv').

    Returns:
        pandas.DataFrame: DataFrame with CSV data (all columns as strings).
    """

    file_location = (
        f"{INPUT_LOCATION}/{book_name}/{sheet_name}/{tag_name}/{csv_name}.csv"
    )
    # must read as string in order to pull through full precision
    df = pd.read_csv(file_location, dtype=str)
    return df


def strip_headers_from_tiny_df(df):
    """
    Special handling for tables with one value.
    XL2TIMES outputs these with a header called VALUE, but we
    need the value to be in the header column (with no data underneath).
    This replaces the header with the value for specific tables.
    """
    df = df.T  # Transpose the df
    df.columns = [df.iloc[0][0]]  # Set the column name to the value
    df = df.iloc[0:0]  # Remove all rows
    return df


def return_csvs_in_folder(folder_name):
    """
    Returns a list of base names for all CSV files in the folder.
    Args:
        folder_name (str): Path to the folder to search for CSV files.
    Returns:
        list of str: Base names of all CSV files found in the folder.
    """
    path = os.path.abspath(folder_name)
    base_names = [
        os.path.splitext(f)[0] for f in os.listdir(path) if f.lower().endswith(".csv")
    ]
    return base_names


def return_subfolders(folder_name):
    """
    Returns a list of names of all subfolders within the folder.
    Used for returning all the tags in a given sheet's folder.
    Args:
        folder_name (str): Path to the folder whose subfolders are listed.
    Returns:
        list: Subfolder names (str) within the specified folder.
    Raises:
        FileNotFoundError: If the folder does not exist.
        NotADirectoryError: If the path is not a directory.
    """
    path = os.path.abspath(folder_name)
    subfolder_names = [f.name for f in os.scandir(path) if f.is_dir()]
    return subfolder_names


def get_sheets_for_book(book_name):
    """
    Retrieves the list of sheet subfolders for a given book.
    Args:
        book_name (str): Name of the book whose sheets are retrieved.
    Returns:
        list: Subfolder names representing sheets in the book folder.
    """
    book_folder = f"{INPUT_LOCATION}/{book_name}"
    sheets = return_subfolders(book_folder)
    return sheets


def get_tags_for_sheet(book_name, sheet_name):
    """
    Retrieves the list of tags (subfolder names) for a sheet in a workbook.
    Args:
        book_name (str): Name of the workbook.
        sheet_name (str): Name of the sheet in the workbook.
    Returns:
        list: Subfolder names (tags) found in the sheet's folder.
    """
    sheet_folder = f"{INPUT_LOCATION}/{book_name}/{sheet_name}"
    tags = return_subfolders(sheet_folder)
    return tags


def get_metadata_df():
    """
    Reads the metadata CSV file, processes 'tag_counter' to generate
    a 'csv_name' column (e.g., 'data_a', 'data_b', ...), and returns the DataFrame.
    Returns:
        pandas.DataFrame: Metadata DataFrame with an additional 'csv_name' column.
    """
    file_location = f"{INPUT_LOCATION}/metadata.csv"
    df = pd.read_csv(file_location)
    df["csv_name"] = df["tag_counter"].apply(lambda x: string.ascii_lowercase[x - 1])
    df["csv_name"] = df["csv_name"].apply(lambda x: f"data_{x}")
    return df


def get_uc_sets(book_name, sheet_name, tag, csv_name):
    """
    Gets 'uc_sets' metadata for book, sheet, tag, and CSV name.
    Args:
        book_name (str): Book/folder name.
        sheet_name (str): Sheet name.
        tag (str): Tag name.
        csv_name (str): CSV file name.
    Returns:
        dict: Parsed 'uc_sets' from metadata, or empty dict if not found/NaN.
    Raises:
        IndexError: If no matching metadata entry.
        ValueError: If 'uc_sets' can't be parsed.
    Notes:
        Logs warning if multiple metadata entries match.
    """
    metadata = get_metadata_df()

    metadata = metadata[
        (metadata["folder_name"] == book_name)
        & (metadata["sheet_name"] == sheet_name)
        & (metadata["tag_name"] == tag)
        & (metadata["csv_name"] == csv_name)
    ]
    if len(metadata) > 1:
        logger.warning(
            "Warning: metadata filter returned multiple entries. Please review"
        )
    # first row uc_sets (should only be one row)
    uc_set = metadata.iloc[0]["uc_sets"]

    if pd.isna(uc_set):
        uc_set = {}
    else:
        uc_set = literal_eval(uc_set)
    return uc_set


def create_empty_workbook(book_name, sheets, suffix="_test_automate"):
    """
    Creates an empty Excel workbook with specified sheet names and saves it.
    The workbook is saved as `book_name` plus optional `suffix` in OUTPUT_LOCATION.
    If the directory does not exist, it is created. Only the listed sheets are added.
    Args:
        book_name (str): Base name for the workbook file.
        sheets (list of str): List of sheet names to create.
        suffix (str, optional): Suffix for the workbook filename.
            Defaults to "_test_automate".
    Returns:
        None
    """

    # This function creates the workbook with empty sheets
    # Later, data is appended to these sheets by overlay.
    book_location = f"{OUTPUT_LOCATION}/{book_name}{suffix}.xlsx"
    book_directory = os.path.dirname(book_location)

    # create the folder if needed
    os.makedirs(book_directory, exist_ok=True)

    # create workbook:
    wb = Workbook()

    # remove the default sheet which gets activated
    wb.remove(wb.active)

    # add each sheet
    for sheet in sheets:
        wb.create_sheet(sheet)

    # save
    wb.save(book_location)


def write_data(df, book_name, sheet_name, tag, uc_set, startrow=0):
    """
    Writes a DataFrame to an Excel workbook/sheet,
        including custom tags and UC_Set metadata.
    Parameters:
        df (pandas.DataFrame): Data to write to the Excel sheet.
        book_name (str): Excel workbook name (without extension).
        sheet_name (str): Sheet name in the workbook.
        tag (str): Tag string written above the data table (Veda format).
        uc_set (dict): Optional UC_Set tags (dict of key/value strings).
        startrow (int, optional): Starting row index for writing (default 0).
    Side Effects:
        Modifies and saves the Excel workbook by writing tag, UC_Set tags, headers,
        and data.
    Notes:
        - Workbook must exist at OUTPUT_LOCATION.
        - Tag is formatted and written above the data table.
        - UC_Set tags (if any) are written above the tag row.
        - Data and headers are written starting from the calculated row index.
    """

    # This function is quite big. However, it's not part of the main workflow
    # So I will not refactor it to ensure it passes function size rules.
    # Consider it a historical relic and sorry for all the locals in it

    # pylint: disable=too-many-locals
    # pylint: disable=too-many-arguments
    # pylint: disable=too-many-positional-arguments

    new_workbook = f"{OUTPUT_LOCATION}/{book_name}.xlsx"

    # Fix up the tag to match Veda expectations
    tag = f"~{tag}"
    tag = tag.replace("·", ":")

    # Get uc_set length
    uc_set_length = len(uc_set)
    if uc_set_length > 0:
        startrow += uc_set_length - 1

    # Load existing workbook
    book = load_workbook(new_workbook)
    sheet = book[sheet_name]

    # Write the header row
    for col_idx, column_name in enumerate(df.columns, 1):
        sheet.cell(row=startrow + 2, column=col_idx, value=column_name)

    # Write the data
    for row_idx, row in enumerate(df.values, startrow + 3):
        for col_idx, value in enumerate(row, 1):
            sheet.cell(row=row_idx, column=col_idx, value=value)

    # Write the tag
    tag_row = startrow + 1
    sheet.cell(row=tag_row, column=1, value=tag)

    # Add UC_Set tags if needed
    if uc_set_length > 0:
        # first move the table tag to B
        sheet.cell(row=tag_row, column=2, value=tag)
        for n in range(uc_set_length):
            uc_set_tag_row = startrow - n + 1
            key = list(uc_set.keys())[n]
            value = uc_set[key]
            sheet.cell(row=uc_set_tag_row, column=1, value=f"~UC_Sets: {key}: {value}")

    # Save the workbook
    book.save(new_workbook)


def write_all_tags_to_sheet(book_name, sheet_name):
    """
    This function writes all tags for a given sheet in a book.
    It retrieves the tags for the sheet, reads the corresponding CSV files,
    and writes the data to the sheet in the workbook.

    It handles multiple tags by stacking them vertically
    and ensures that the data does not overwrite previous entries.
    It also manages the row indices to ensure that each tag's data
    is written in the correct position without overwriting other tags.

    :param book_name: The name of the book for which to write the tags.
    :param sheet_name: The name of the sheet to which the tags will be written.
    :return: None
    """

    # The sheets with multiple tags need to be stacked up.
    # Each table we write will need to have its row number
    # saved in here so other tables can move down.

    tag_list = get_tags_for_sheet(book_name, sheet_name)
    sheet_folder = f"{INPUT_LOCATION}/{book_name}/{sheet_name}"

    # we start from the first row and will move down as needed
    startrow = 0

    for tag_name in tag_list:
        csv_files = return_csvs_in_folder(f"{sheet_folder}/{tag_name}")

        for csv_name in csv_files:
            # read the data
            df = get_csv_data(book_name, sheet_name, tag_name, csv_name)
            # include the uc_set if needed (this comes up null otherwise)
            uc_set = get_uc_sets(book_name, sheet_name, tag_name, csv_name)

            # Patch for small files
            # TO-DO: automate handling of these rather than
            # hardcoding which tables receive this treatment.
            # This currently covers all of TIMES-NZ but better
            # to be flexible to future changes if needed.

            if (
                book_name == "SysSettings"
                and sheet_name == "TimePeriods"
                and tag_name in ["StartYear", "ActivePDef"]
            ):
                df = strip_headers_from_tiny_df(df)

            # create the tag (this also returns the colons where necessary)
            write_data(df, book_name, sheet_name, tag_name, uc_set, startrow=startrow)
            # measure the length (row count), adding extra
            # space for additional uc_sets if needed, so the
            # next table has space
            df_row_count = len(df) + len(uc_set)
            # add the dataframe rows to our start row index
            # so we can keep going without overwriting
            # and additional rows for a healthy gap.
            startrow += df_row_count + 3


def write_workbook(book_name):
    """
    This function creates a new workbook for the given book_name,
    creates the sheets based on the book's structure,
    and writes all tags to each sheet.
    It overwrites any existing workbook with the same name.

    :param book_name: The name of the book for which to create the workbook.
    :return: None
    """
    logger.info("Creating %s.xlsx:", book_name)
    sheets = get_sheets_for_book(book_name)
    # create structure, overwriting everything already there
    create_empty_workbook(book_name, sheets, suffix="")

    for sheet in sheets:
        # Verbose printing
        logger.info("     - Sheet: '%s'", sheet)
        # the workbook exists now we write each tag set to each sheet
        write_all_tags_to_sheet(book_name, sheet_name=sheet)
