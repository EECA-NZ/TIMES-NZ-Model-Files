"""Tests for agriculture demand projection helpers."""

import csv
import sys
import types
from pathlib import Path

import pandas as pd
from openpyxl import Workbook

fake_stage_0_settings = types.ModuleType("prepare_times_nz.stage_0.stage_0_settings")
fake_stage_0_settings.BASE_YEAR = 2023
sys.modules.setdefault(
    "prepare_times_nz.stage_0.stage_0_settings", fake_stage_0_settings
)

# pylint: disable=wrong-import-position
from prepare_times_nz.stage_3.demand_projections import agriculture

# pylint: disable= duplicate-code
ASSUMPTION_COLUMNS = [
    "SectorGroup",
    "Sector",
    "Scenario",
    "Method",
    "Workbook",
    "SheetName",
    "SourceCategory1",
    "SourceCategory2",
    "ConstantIndex",
    "Note",
]

WORKBOOK_NAME = "mfe/Detailed-results-for-ERP2-projection-scenarios.xlsx"
SECTOR_GROUP = "Agriculture, Forestry and Fishing"

TEST_ASSUMPTIONS = [
    {
        "SectorGroup": SECTOR_GROUP,
        "Sector": "Dairy Cattle Farming",
        "Scenario": "Traditional",
        "Method": "Workbook",
        "Workbook": WORKBOOK_NAME,
        "SheetName": "Baseline",
        "SourceCategory1": "",
        "SourceCategory2": "Total dairy cattle",
        "ConstantIndex": "",
        "Note": "",
    },
    {
        "SectorGroup": SECTOR_GROUP,
        "Sector": "Dairy Cattle Farming",
        "Scenario": "Transformation",
        "Method": "Workbook",
        "Workbook": WORKBOOK_NAME,
        "SheetName": "Baseline low",
        "SourceCategory1": "",
        "SourceCategory2": "Total dairy cattle",
        "ConstantIndex": "",
        "Note": "",
    },
    {
        "SectorGroup": SECTOR_GROUP,
        "Sector": "Livestock Farming",
        "Scenario": "Traditional",
        "Method": "Workbook",
        "Workbook": WORKBOOK_NAME,
        "SheetName": "Baseline",
        "SourceCategory1": "",
        "SourceCategory2": "Sheep and beef 'stock units'",
        "ConstantIndex": "",
        "Note": "",
    },
    {
        "SectorGroup": SECTOR_GROUP,
        "Sector": "Forestry and Logging",
        "Scenario": "Traditional",
        "Method": "Workbook",
        "Workbook": WORKBOOK_NAME,
        "SheetName": "Baseline",
        "SourceCategory1": "Forestry (million m3)",
        "SourceCategory2": "Harvested timber (TRV)",
        "ConstantIndex": "",
        "Note": "",
    },
    {
        "SectorGroup": SECTOR_GROUP,
        "Sector": "Indoor Cropping",
        "Scenario": "Traditional",
        "Method": "Workbook",
        "Workbook": WORKBOOK_NAME,
        "SheetName": "Baseline",
        "SourceCategory1": "",
        "SourceCategory2": "Horticulture",
        "ConstantIndex": "",
        "Note": "",
    },
    {
        "SectorGroup": SECTOR_GROUP,
        "Sector": "Fishing, Hunting and Trapping",
        "Scenario": "Traditional",
        "Method": "Constant",
        "Workbook": "",
        "SheetName": "",
        "SourceCategory1": "",
        "SourceCategory2": "",
        "ConstantIndex": 1,
        "Note": "",
    },
]


def make_test_erp_workbook(path: Path):
    """Create a minimal ERP-like workbook with scenario sheets and year columns."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Baseline"

    for sheet_name, dairy_2050, livestock_2050 in [
        ("Baseline", 70, 80),
        ("Baseline low", 60, 65),
    ]:
        ws = (
            wb[sheet_name]
            if sheet_name in wb.sheetnames
            else wb.create_sheet(sheet_name)
        )
        ws.append([None, None])
        ws.append(["Scenario", None])
        ws.append([sheet_name, None])
        ws.append([None, None])
        ws.append([None, None])
        ws.append([None, None, 2023, 2025, 2030, 2050])
        ws.append([None, "Total dairy cattle", 100, 90, 80, dairy_2050])
        ws.append([None, "Sheep and beef 'stock units'", 100, 95, 90, livestock_2050])
        ws.append(
            ["Forestry (million m3)", "Harvested timber (TRV)", 100, 110, 130, 160]
        )
        ws.append([None, "Horticulture", 100, 105, 110, 120])
        ws.append(["Other agriculture", "Total", 100, 102, 104, 106])

    wb.save(path)


def write_assumptions_csv(path: Path, rows):
    """Write the agriculture assumptions test file from structured rows."""
    with path.open("w", newline="", encoding="utf-8") as file_obj:
        writer = csv.DictWriter(file_obj, fieldnames=ASSUMPTION_COLUMNS)
        writer.writeheader()
        writer.writerows(rows)


def write_baseyear_demand_csv(path: Path):
    """Write a minimal base-year demand file for agriculture projection tests."""
    rows = [
        {
            "Sector": "Dairy Cattle Farming",
            "CommodityOut": "MILK",
            "Island": "NI",
            "Technology": "DAIRY_TECH",
            "EndUse": "Process",
            "Variable": "InputEnergy",
            "Unit": "PJ",
            "Value": 10.0,
        },
        {
            "Sector": "Dairy Cattle Farming",
            "CommodityOut": "MILK",
            "Island": "NI",
            "Technology": "DAIRY_TECH",
            "EndUse": "Process",
            "Variable": "OutputEnergy",
            "Unit": "PJ",
            "Value": 5.0,
        },
        {
            "Sector": "Fishing, Hunting and Trapping",
            "CommodityOut": "FISH",
            "Island": "SI",
            "Technology": "FISH_TECH",
            "EndUse": "Process",
            "Variable": "InputEnergy",
            "Unit": "PJ",
            "Value": 3.0,
        },
    ]
    path.parent.mkdir(parents=True, exist_ok=True)
    pd.DataFrame(rows).to_csv(path, index=False)


def setup_temp_agriculture_paths(tmp_path: Path, monkeypatch):
    """Point the agriculture module at temporary stage data locations."""
    stage_2_data = tmp_path / "stage_2_baseyear_data"
    stage_3_data = tmp_path / "stage_3_scenario_data"
    output = stage_3_data / "demand_projections"

    monkeypatch.setattr(agriculture, "STAGE_2_DATA", stage_2_data)
    monkeypatch.setattr(agriculture, "STAGE_3_DATA", stage_3_data)
    monkeypatch.setattr(agriculture, "OUTPUT", output)
    monkeypatch.setattr(agriculture, "OUTPUT_CHECKS", output / "checks")

    return stage_2_data, stage_3_data


def get_index(df, sector, scenario, year):
    """Return one compiled index value for the requested sector/scenario/year."""
    return df[
        (df["Sector"] == sector) & (df["Scenario"] == scenario) & (df["Year"] == year)
    ]["Index"].iloc[0]


def test_get_agriculture_growth_indices_reads_workbook_mappings(tmp_path):
    """Workbook-mapped and constant agriculture projections should compile correctly."""
    external_data_dir = tmp_path / "external_data"
    workbook_dir = external_data_dir / "mfe"
    workbook_dir.mkdir(parents=True)
    workbook_path = workbook_dir / "Detailed-results-for-ERP2-projection-scenarios.xlsx"
    make_test_erp_workbook(workbook_path)

    assumptions_path = tmp_path / "agriculture_demand_projections.csv"
    write_assumptions_csv(assumptions_path, TEST_ASSUMPTIONS)

    df = agriculture.get_agriculture_growth_indices(
        assumptions_path=assumptions_path,
        external_data_dir=external_data_dir,
    )

    assert get_index(df, "Dairy Cattle Farming", "Traditional", 2025) == 0.9
    assert get_index(df, "Dairy Cattle Farming", "Traditional", 2024) == 0.95
    assert get_index(df, "Dairy Cattle Farming", "Transformation", 2050) == 0.6
    assert get_index(df, "Livestock Farming", "Traditional", 2050) == 0.8
    assert get_index(df, "Indoor Cropping", "Traditional", 2050) == 1.2
    assert get_index(df, "Fishing, Hunting and Trapping", "Traditional", 2042) == 1.0


def test_get_energy_demand_projections_uses_temp_stage_data(tmp_path, monkeypatch):
    """Energy projections should be computed from temp test data, not repo intermediates."""
    stage_2_data, _ = setup_temp_agriculture_paths(tmp_path, monkeypatch)

    external_data_dir = tmp_path / "external_data"
    workbook_dir = external_data_dir / "mfe"
    workbook_dir.mkdir(parents=True)
    workbook_path = workbook_dir / "Detailed-results-for-ERP2-projection-scenarios.xlsx"
    make_test_erp_workbook(workbook_path)

    assumptions_path = tmp_path / "agriculture_demand_projections.csv"
    write_assumptions_csv(assumptions_path, TEST_ASSUMPTIONS)
    monkeypatch.setattr(agriculture, "ASSUMPTIONS_FILE", assumptions_path)
    monkeypatch.setattr(agriculture, "EXTERNAL_DATA", external_data_dir)
    baseyear_path = (
        stage_2_data / "ag_forest_fish" / "baseyear_ag_forest_fish_demand.csv"
    )

    write_baseyear_demand_csv(baseyear_path)

    df = agriculture.get_energy_demand_projections("InputEnergy")

    dairy_2025 = df[
        (df["Sector"] == "Dairy Cattle Farming")
        & (df["Scenario"] == "Traditional")
        & (df["Year"] == 2025)
        & (df["Variable"] == "InputEnergy")
    ]["Value"].iloc[0]
    fishing_2042 = df[
        (df["Sector"] == "Fishing, Hunting and Trapping")
        & (df["Scenario"] == "Traditional")
        & (df["Year"] == 2042)
        & (df["Variable"] == "InputEnergy")
    ]["Value"].iloc[0]

    assert dairy_2025 == 9.0
    assert fishing_2042 == 3.0


def test_main_writes_outputs_to_temp_stage_data(tmp_path, monkeypatch):
    """Main should save outputs under temp stage data rather than repo intermediates."""
    stage_2_data, stage_3_data = setup_temp_agriculture_paths(tmp_path, monkeypatch)

    external_data_dir = tmp_path / "external_data"
    workbook_dir = external_data_dir / "mfe"
    workbook_dir.mkdir(parents=True)
    workbook_path = workbook_dir / "Detailed-results-for-ERP2-projection-scenarios.xlsx"
    make_test_erp_workbook(workbook_path)

    assumptions_path = tmp_path / "agriculture_demand_projections.csv"
    write_assumptions_csv(assumptions_path, TEST_ASSUMPTIONS)
    monkeypatch.setattr(agriculture, "ASSUMPTIONS_FILE", assumptions_path)
    monkeypatch.setattr(agriculture, "EXTERNAL_DATA", external_data_dir)
    baseyear_path = (
        stage_2_data / "ag_forest_fish" / "baseyear_ag_forest_fish_demand.csv"
    )

    write_baseyear_demand_csv(baseyear_path)

    agriculture.main()

    assert (
        stage_3_data / "demand_projections" / "agriculture_demand_index.csv"
    ).exists()
    assert (
        stage_3_data / "demand_projections" / "checks" / "agriculture_input.csv"
    ).exists()
    assert (
        stage_3_data / "demand_projections" / "checks" / "agriculture_output.csv"
    ).exists()
